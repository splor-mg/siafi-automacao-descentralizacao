import os
import sys
import shutil
from py3270 import Emulator
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
from dotenv import load_dotenv
from fluxo_anular_desc import anulacao
from fluxo_aprovar_desc import aprovacao

# Carrega as credenciais do arquivo .env (na raiz do repositório)
load_dotenv()

sistema = 'simg'
usuario           = os.getenv('USUARIO')
senha             = os.getenv('SENHA')
unidade_executora = os.getenv('UNIDADE_EXECUTORA')
unidade_orcamentaria = os.getenv('UNIDADE_ORCAMENTARIA')
month = datetime.today().strftime("%m")

# Definição dos CAMINHOS
# Raiz do repositório, derivada da localização deste script
# (siafi_automacao/descentralizacao.py -> sobe um nível).
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Pasta de trabalho local (Linux). Vem do .env (PASTA_LINUX); se ausente,
# usa a pasta 'data' na raiz do repositório.
DATA = os.getenv('PASTA_LINUX') or os.path.join(BASE_DIR, 'data')

# Pasta-raiz do projeto no Windows/OneDrive (via /mnt/c). Vem do .env
# (PASTA_WINDOWS). A subpasta "Conferencia" guarda o arquivo gerado pelo
# consolida.py, que é trazido para cá, processado, e devolvido ao final.
PASTA_WINDOWS = os.getenv('PASTA_WINDOWS')
SUB_CONFERENCIA = 'Conferencia'

# CAMINHO_CONSOLIDADO é definido em tempo de execução por
# preparar_arquivo_para_processar() (o nome do arquivo varia por dia).
CAMINHO_CONSOLIDADO = None

# Nome da aba na planilha consolidada onde estão os dados a serem processados
SHEET_NAME = 'Execucao'


def preparar_arquivo_para_processar():
    """Traz o arquivo de conferência (gerado pelo consolida.py) da pasta
    Conferencia (Windows/OneDrive) para a pasta local de trabalho (DATA) e
    devolve (caminho_local, nome_arquivo). Trabalhar no filesystem Linux torna
    os salvamentos de progresso rápidos e evita locks do OneDrive.
    Em modo local (sem PASTA_WINDOWS) usa DATA/consolidado.xlsx."""
    os.makedirs(DATA, exist_ok=True)

    if not PASTA_WINDOWS:
        return os.path.join(DATA, 'consolidado.xlsx'), None

    conf_dir = os.path.join(PASTA_WINDOWS, SUB_CONFERENCIA)
    xlsxs = []
    if os.path.isdir(conf_dir):
        xlsxs = [f for f in os.listdir(conf_dir)
                 if f.lower().endswith('.xlsx') and not f.startswith('~$')]
    if not xlsxs:
        raise FileNotFoundError(
            f'Nenhum arquivo de conferência em "{conf_dir}". '
            f'Rode o consolida.py antes do descentralizacao.py.')

    # Mais recente (caso haja mais de um do dia)
    xlsxs.sort(key=lambda f: os.path.getmtime(os.path.join(conf_dir, f)))
    nome = xlsxs[-1]
    destino_local = os.path.join(DATA, nome)
    shutil.move(os.path.join(conf_dir, nome), destino_local)
    print(f'Arquivo de conferência trazido para processamento local: {nome}')
    return destino_local, nome


def devolver_arquivo_processado(caminho_local, nome):
    """Devolve o arquivo já processado para a pasta Conferencia (Windows)."""
    if not (PASTA_WINDOWS and nome):
        return
    conf_dir = os.path.join(PASTA_WINDOWS, SUB_CONFERENCIA)
    os.makedirs(conf_dir, exist_ok=True)
    destino = os.path.join(conf_dir, nome)
    shutil.move(caminho_local, destino)
    print(f'Arquivo processado devolvido para: {destino}')


def salvar_progresso(df, caminho, sheet):
    """Salva o DataFrame de volta na planilha de forma atômica, preservando
    o que já foi gravado na coluna 'Progresso'."""
    pasta = os.path.dirname(caminho)
    tmp = os.path.join(pasta, '_tmp_consolidado.xlsx')
    with pd.ExcelWriter(tmp, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet, index=False)
    os.replace(tmp, caminho)


def traduzir_progresso(retorno):
    """Converte a mensagem crua do SIAFI no texto que vai para a coluna
    'Progresso'. Sucesso ('REGISTRO EFETUADO') vira 'Ok'; mensagens de erro
    conhecidas viram um texto amigável; qualquer outro retorno é mantido como
    veio, para nada ser escondido. Para incluir novas mensagens, basta
    acrescentar uma linha no mapa abaixo."""
    if retorno is None or (isinstance(retorno, str) and retorno.strip() == ''):
        return ''
    retorno = retorno.strip()

    # Sucesso
    if 'REGISTRO EFETUADO' in retorno.upper():
        return 'Ok'

    if retorno.startswith("E90 - SALDO ZERADO NA CONTA"):
        return 'Saldo zerado na conta'

    mapa = {
        "0139- PROJ/ATIV OU FONTE/PROC./IAG INEXISTENTE PARA UO":
            'Proj/Ativ ou Fonte/Proc./IAG inexistente para a UO',
        "0101- NATUREZA DESPESA INEXISTENTE(S).":
            'Natureza de despesa inexistente',
        "0101- GRUPO DESPESA INEXISTENTE(S).":
            'Grupo de despesa inexistente',
        "0139- PROGRAMA DE TRABALHO NAO ENCONTRADO PARA GM/FP.":
            'Programa de trabalho não encontrado para GM/FP',
        "0109-INFORME NATUREZA DE DESPESA COMPLETA.":
            'Informe a natureza de despesa completa',
        "0139- NAO EXITE DESCENTRALIZACAO PARA PROJ/ATIV OU FONTE/PROC./IAG":
            'Não existe descentralização para Proj/Ativ ou Fonte/Proc./IAG',
        "0139- SALDO INEXISTENTE A ANULAR PARA PROJ/ATIV":
            'Saldo inexistente a anular para Proj/Ativ',
        "0139- VALOR A APROVAR MAIOR QUE SALDO DISPONIVEL NO PROJ/ATIV.":
            'Valor a aprovar maior que o saldo disponível',
        "SALDO DE CREDITO ORCAMENTARIO A APROVAR POR PROJ/ATIV ZERADO.":
            'Saldo de crédito a aprovar zerado',
        "0139- VALOR A DESCENTRALIZAR MAIOR QUE SALDO APROVADO.":
            'Valor a descentralizar maior que o saldo aprovado',
        "0139- VALOR INFORMADO ALEM DO VALOR DESCENTRALIZADO PARA O PROJETO/ATIVIDADE":
            'Valor informado além do valor descentralizado',
    }
    return mapa.get(retorno, retorno)


def formatar_planilha(ws):
    """Aplica formatação visual gerencial na aba: cabeçalho colorido, coluna
    'Valor' no formato xxx.xxx,xx, zebra nas linhas, coluna 'Progresso' com cor
    condicional (Ok = verde, erros = vermelho/amarelo) e larguras ajustadas."""
    AZUL_ESCURO  = PatternFill('solid', fgColor='1F4E79')
    AZUL_CLARO   = PatternFill('solid', fgColor='D6E4F0')
    BRANCO       = PatternFill('solid', fgColor='FFFFFF')
    VERDE        = PatternFill('solid', fgColor='C6EFCE')
    AMARELO      = PatternFill('solid', fgColor='FFEB9C')
    VERMELHO     = PatternFill('solid', fgColor='FFC7CE')
    FONTE_BRANCA = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    FONTE_NORMAL = Font(name='Calibri', size=10)
    BORDA = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF'),
    )
    COLS_VALOR = {'Valor'}

    max_col = ws.max_column
    max_row = ws.max_row

    # Mapeia nome da coluna -> índice
    cabec = {ws.cell(row=1, column=c).value: c for c in range(1, max_col + 1)}

    # Remove linhas de grade
    ws.sheet_view.showGridLines = False

    # Cabeçalho
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill   = AZUL_ESCURO
        cell.font   = FONTE_BRANCA
        cell.border = BORDA
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
    ws.row_dimensions[1].height = 20

    # Linhas de dados
    for r in range(2, max_row + 1):
        zebra = AZUL_CLARO if r % 2 == 0 else BRANCO
        for c in range(1, max_col + 1):
            cell      = ws.cell(row=r, column=c)
            cell.fill = zebra
            cell.font = FONTE_NORMAL
            cell.border = BORDA

            col_nome = ws.cell(row=1, column=c).value

            # Formata a coluna de valor monetário no padrão xxx.xxx,xx
            if col_nome in COLS_VALOR and cell.value not in (None, ''):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Cor condicional na coluna Progresso
        if 'Progresso' in cabec:
            prog_cell = ws.cell(row=r, column=cabec['Progresso'])
            valor = (prog_cell.value or '').strip()
            if valor == 'Ok':
                prog_cell.fill = VERDE
            elif valor != '':
                prog_cell.fill = VERMELHO if 'zerado' in valor.lower() or 'maior' in valor.lower() or 'inexistente' in valor.lower() else AMARELO

    # Congela a primeira linha
    ws.freeze_panes = 'A2'

    # Ajusta largura: usa o maior entre o título do cabeçalho e o conteúdo
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for r in range(1, max_row + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 50)


em = Emulator(visible=True) ##caso queira que a tela apareça utilize visible=True
em.connect('bhmvsb.prodemge.gov.br')
em.wait_for_field()

# Preenche os dados de login
em.fill_field(19, 13, sistema, 7)
em.fill_field(20, 13, usuario, 8)
em.fill_field(21, 13, senha, 7)
em.send_enter()

# Loop: navega pelas telas até encontrar a mensagem de sucesso
max_tentativas = 10
tentativas = 0

while tentativas < max_tentativas:
    time.sleep(1)

    try:
        em.send_enter()

        # Tela COM campo editável — verifica se é a tela de sucesso
        if em.string_found(1, 13, 'Logon executado com sucesso'):
            print("Login realizado com sucesso!")
            break
        elif em.string_found(23, 2, 'Senha expirada. Digite Nova Senha.'):
            print("")
            print("=" * 70)
            print("Senha expirada. Abra o SIAFI manualmente e atualize sua senha.")
            print("Apos isso, salve a nova senha no arquivo .env e execute o script novamente.")
            print("=" * 70)
            em.terminate()
            sys.exit(3)
            break
        else:
            # Tela com campo editável, mas ainda não é a de sucesso
            print(f"Tentativa {tentativas + 1} - tela intermediária, avançando...")
            em.send_enter()

    except Exception:
        print(f"Tentativa {tentativas + 1} - tela de aviso detectada, passando...")
        em.send_enter()

    tentativas += 1

if tentativas == max_tentativas:
    print("Não foi possível fazer login após várias tentativas.")
    em.terminate()

em.fill_field(1, 2, sistema, 4)
em.send_enter()

##nova tela buscando login...
max_tentativas = 10
tentativas = 0

while tentativas < max_tentativas:
    time.sleep(1)

    try:
        em.send_enter()

        # Tela COM campo editável — verifica se é a tela de sucesso
        if em.string_found(22, 11, 'Unidade Executora'):
            print("Texto encontrado")
            break

        else:
            # Tela com campo editável, mas ainda não é a de sucesso
            print(f"Tentativa {tentativas + 1} - tela intermediária, avançando...")
            em.send_enter()

    except:
        # Tela SEM campo editável — é a tela de aviso, só dá Enter e segue
        print(f"Tentativa {tentativas + 1} - tela de aviso detectada, passando...")
        em.send_enter()

    tentativas += 1

if tentativas == max_tentativas:
    print("Não foi possível fazer login após várias tentativas.")
    em.terminate()

#Entrar com a Unidade Executora
em.fill_field(22, 30, unidade_executora, 7)
em.send_enter()
em.wait_for_field()
# Fim do login

#Entrar em 03 - Movimentacao Orcamentaria
em.fill_field(21, 19, '03', 2)
em.send_enter()
em.wait_for_field()

#Entrar em 03 - Descentralizacao de Cota Orcamentaria
em.fill_field(21, 19, '03', 2)
em.send_enter()
em.wait_for_field()

# Leitura da planilha e processamento dos dados

# -----------------------------------------------------------------------
# ETAPA 3: leitura da planilha consolidada
# -----------------------------------------------------------------------
# Traz o arquivo de conferência do Windows/OneDrive para a pasta local
CAMINHO_CONSOLIDADO, _nome_conferencia = preparar_arquivo_para_processar()

df = pd.read_excel(CAMINHO_CONSOLIDADO, sheet_name=SHEET_NAME)
df = df.dropna(how='all')  # remove linhas completamente vazias
df = df.reset_index(drop=True)

# Garante que a coluna 'Progresso' exista e aceite texto (retorno do SIAFI)
if 'Progresso' not in df.columns:
    df['Progresso'] = pd.NA
df['Progresso'] = df['Progresso'].astype('object')

# O loop usa "for idx, row" para que o df.at[idx, 'Progresso'] grave o retorno
# na linha correta em memória (e depois em disco).
puladas = 0
for idx, row in df.iterrows():
    # Linhas com 'Progresso' já preenchido foram executadas no SIAFI em uma
    # rodada anterior. Pular é o que torna seguro reexecutar o robô depois de
    # uma queda de conexão: sem isso, ele refaria as operações já realizadas.
    progresso_atual = row['Progresso']
    if pd.notna(progresso_atual) and str(progresso_atual).strip() != '':
        print(f"Linha {idx + 2}: já processada ('{str(progresso_atual).strip()}') — pulando.")
        puladas += 1
        continue

    data_row = {}
    data_row['month']   = month
    data_row['orientacao']      = str((row['Orientacao']))
    data_row['ue']   = str(int(row['UE_Beneficiada']))
    data_row['tipo']     = str((row['Tipo de Descentralizacao']))
    data_row['fonte']   = str(int(row['Fonte']))
    data_row['procedencia'] = str(int(row['Procedencia']))
    data_row['dea'] = str((row['Elemento 92']))
    data_row['acao'] = str(int(row['Acao']))
    data_row['natureza_despesa'] = str(int(row['Natureza_Despesa_Elemento']))

    if pd.notna(row['Natureza_Despesa_Elemento']):
        data_row['categoria'] = data_row['natureza_despesa'][:1]   # dois primeiros digitos
        data_row['grupo'] = data_row['natureza_despesa'][1:2]       # segundo digito
        data_row['modalidade'] = data_row['natureza_despesa'][2:4]  # terceiro
        data_row['elemento'] = data_row['natureza_despesa'][4:6]    # dois ultimos digitos
    else:
        data_row['categoria'] = '0'
        data_row['grupo'] = '0'
        data_row['modalidade'] = '0'
        data_row['elemento'] = '0'

    data_row['item'] = str(int(row['Item'])) if pd.notna(row['Item']) else '0'
    data_row['uo_financiadora'] = str(int(row['UO_Financiadora'])) if pd.notna(row['UO_Financiadora']) else '0'
    data_row['iag'] = str(int(row['IAG']))
    data_row['unidade_orcamentaria'] = unidade_orcamentaria  # UO da UE executora (vem do .env)
    data_row['valor'] = int(round(float(row['Valor']) * 100))

    # Criação da Variável de Retorno para armazenar o resultado do processamento de cada linha
    retorno = ''

    if data_row['orientacao'] == 'Anular':
        print(f"Realizando procedimento de anulação")
    elif data_row['orientacao'] == 'Aprovar':
        print(f"Realizando procedimento de aprovação")

    if data_row['tipo'] == 'Global':        
        print(f" UE = {data_row['ue']}, Tipo = {data_row['tipo']}, Fonte e IPU = {data_row['fonte']}.{data_row['procedencia']}, UO Financiadora = {data_row['uo_financiadora']}, Ação = {data_row['acao']}, Natureza de Despesa = {data_row['natureza_despesa']}, Valor: {data_row['valor']}")
    elif data_row['tipo'] == 'Amarrado':
        print(f" UE = {data_row['ue']}, Tipo = {data_row['tipo']}, Fonte e IPU = {data_row['fonte']}.{data_row['procedencia']}, Uo Financiadora = {data_row['uo_financiadora']}, Ação = {data_row['acao']}, Natureza de Despesa = {data_row['natureza_despesa']}{data_row['item']}, Valor: {data_row['valor']}")


    # -------------------- exemplo para orquestrar o fluxo --------------------
    # aqui você pode inspecionar o data_row e decidir se é anulação ou aprovação, global ou amarrado, e então chamar as funções correspondentes

    if data_row['orientacao'] == 'Anular':
        retorno = anulacao(em, data_row)
    elif data_row['orientacao'] == 'Aprovar':
        retorno = aprovacao(em, data_row)

    # Traduz o retorno do SIAFI e grava na coluna 'Progresso', salvando
    # imediatamente para não perder o que já foi processado caso pare no meio.
    df.at[idx, 'Progresso'] = traduzir_progresso(retorno)
    salvar_progresso(df, CAMINHO_CONSOLIDADO, SHEET_NAME)

if puladas:
    print(f"{puladas} linha(s) já processada(s) foram puladas nesta execução.")

# Salvamento final (redundante com o incremental, mas garante o estado completo)
salvar_progresso(df, CAMINHO_CONSOLIDADO, SHEET_NAME)

# Aplica a formatação visual na planilha salva
wb = load_workbook(CAMINHO_CONSOLIDADO)
ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
formatar_planilha(ws)
wb.save(CAMINHO_CONSOLIDADO)
print(f"Progresso gravado e planilha formatada em: {CAMINHO_CONSOLIDADO}")

# Devolve o arquivo já processado para a pasta Conferencia (Windows/OneDrive)
devolver_arquivo_processado(CAMINHO_CONSOLIDADO, _nome_conferencia)

em.terminate()